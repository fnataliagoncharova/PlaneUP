from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
from db import get_connection
import datetime
import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
import re

ALLOWED_UNITS = ["м²", "м.п."]


def normalize_unit(value: Optional[str]) -> str:
    """
    Normalize user-provided unit strings to canonical values.
    Returns one of ALLOWED_UNITS or raises HTTPException 400.
    """
    if value is None:
        raise HTTPException(
            status_code=400,
            detail=f"Недопустимая единица измерения. Допустимые значения: {', '.join(ALLOWED_UNITS)}",
        )

    def clean(s: str) -> str:
        # lower-case and strip spaces/dots for comparison
        return re.sub(r"[\\s\\.]+", "", s.lower())

    cleaned = clean(str(value))

    m2_aliases = [
        "м2",
        "м^2",
        "м²",
        "м 2",
        "кв.м",
        "кв м",
        "квадратный метр",
        "квадратные метры",
        "квадратный",
        "квадратные",
        "м кв",
    ]
    mp_aliases = [
        "м",
        "м.п.",
        "мп",
        "м п",
        "пог.м",
        "пог м",
        "погонный метр",
        "погонные метры",
        "погонный",
        "погонные",
        "пм",
        "мпог",
        "м пог",
    ]

    m2_variants = {clean(a) for a in m2_aliases}
    mp_variants = {clean(a) for a in mp_aliases}

    if cleaned in m2_variants:
        return ALLOWED_UNITS[0]
    if cleaned in mp_variants:
        return ALLOWED_UNITS[1]

    raise HTTPException(
        status_code=400,
        detail=f"Недопустимая единица измерения. Допустимые значения: {', '.join(ALLOWED_UNITS)}",
    )

class ProductUpdate(BaseModel):
    product_name: Optional[str] = None
    unit_of_measure: Optional[str] = None
    is_active: Optional[bool] = None

class SemiFinishedUpdate(BaseModel):
    semi_finished_name: Optional[str] = None
    unit_of_measure: Optional[str] = None
    is_active: Optional[bool] = None
    degas_days: Optional[int] = None


app = FastAPI(title="PlaneUP API")

# Allow opening the HTML prototype from file:// or any local dev host.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*", "null"],  # file:// origin is "null"
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
def root():
    return {"message": "PlaneUP backend is running"}


@app.get("/health/db")
def health_db():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("select current_database();")
    result = cur.fetchone()
    cur.close()
    conn.close()

    return {"database": result[0]}

@app.get("/products")
def get_products():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        select
            product_id,
            product_code,
            product_name,
            coalesce(unit_of_measure, unit) as unit_of_measure,
            coalesce(is_active, active, true) as is_active
        from products
        order by product_code
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    result = []
    for row in rows:
        result.append({
            "product_id": row[0],
            "product_code": row[1],
            "product_name": row[2],
            "unit_of_measure": row[3],
            "is_active": bool(row[4]),
        })

    return result


@app.put("/products/{product_code}")
def update_product(product_code: str, payload: ProductUpdate):
    update_fields = payload.dict(exclude_unset=True)
    if not update_fields:
        return {"status": "no_changes"}

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "select product_id from products where product_code = %s",
            (product_code,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Product not found")
        product_id = row[0]

        set_clauses = []
        params = []

        if "product_name" in update_fields:
            set_clauses.append("product_name = %s")
            params.append(update_fields["product_name"])

        if "unit_of_measure" in update_fields:
            update_fields["unit_of_measure"] = normalize_unit(update_fields["unit_of_measure"])
            # keep legacy unit column in sync
            set_clauses.append("unit_of_measure = %s")
            params.append(update_fields["unit_of_measure"])
            set_clauses.append("unit = %s")
            params.append(update_fields["unit_of_measure"])

        if "is_active" in update_fields:
            set_clauses.append("is_active = %s")
            params.append(update_fields["is_active"])
            set_clauses.append("active = %s")
            params.append(update_fields["is_active"])

        set_clauses.append("updated_at = now()")

        if len(params) == 0:
            return {"status": "no_changes"}

        sql = f"""
            update products
            set {', '.join(set_clauses)}
            where product_id = %s
            returning
                product_id,
                product_code,
                product_name,
                coalesce(unit_of_measure, unit) as unit_of_measure,
                coalesce(is_active, active, true) as is_active
        """
        params.append(product_id)
        cur.execute(sql, tuple(params))
        updated = cur.fetchone()
        conn.commit()

        return {
            "product_id": updated[0],
            "product_code": updated[1],
            "product_name": updated[2],
            "unit_of_measure": updated[3],
            "is_active": bool(updated[4]),
        }
    finally:
        cur.close()
        conn.close()


@app.get("/products/import-template")
def get_products_import_template():
    """
    Downloadable Excel template for bulk import of products.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    headers = ["product_code", "product_name", "unit_of_measure", "is_active"]
    ws.append(headers)
    ws.append(["P-EXAMPLE", "Пример продукции", ALLOWED_UNITS[0], "Да"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="products_import_template.xlsx"'},
    )


@app.post("/products/import")
def import_products(file: UploadFile = File(...)):
    """
    Import products from uploaded Excel file (.xlsx).
    """
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Требуется файл .xlsx")

    try:
        wb = load_workbook(file.file)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать Excel-файл")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Файл не содержит данных")

    header = [str(h).strip() if h else "" for h in rows[0]]
    expected_header = ["product_code", "product_name", "unit_of_measure", "is_active"]
    if [h.lower() for h in header] != expected_header:
        raise HTTPException(status_code=400, detail="Неверный заголовок файла")

    seen_codes = set()
    items = []
    errors = []

    def normalize_bool(val):
        if val is None:
            return None
        if isinstance(val, bool):
            return val
        s = str(val).strip().lower()
        if s in ["да", "yes", "true", "1"]:
            return True
        if s in ["нет", "no", "false", "0"]:
            return False
        return None

    for idx, row in enumerate(rows[1:], start=2):  # 1-based row numbers
        product_code, product_name, unit_of_measure, is_active_raw = row[:4]

        if all(
            (cell is None or (isinstance(cell, str) and cell.strip() == ""))
            for cell in (product_code, product_name, unit_of_measure, is_active_raw)
        ):
            continue  # skip empty row
        row_errors = []

        code = (product_code or "").strip()
        name = (product_name or "").strip()
        try:
            unit = normalize_unit(unit_of_measure)
        except HTTPException as e:
            unit = None
        active_val = normalize_bool(is_active_raw)

        if not code:
            row_errors.append("не заполнен product_code")
        if code in seen_codes:
            row_errors.append("дублирующийся product_code в файле")
        if not name:
            row_errors.append("не заполнен product_name")
        if unit is None:
            row_errors.append("недопустимая единица измерения (разрешено: м² или м.п.)")
        if active_val is None:
            row_errors.append("недопустимое значение is_active (используйте Да/Нет)")

        if row_errors:
            errors.append({"row": idx, "errors": row_errors})
        else:
            seen_codes.add(code)
            items.append(
                {
                    "product_code": code,
                    "product_name": name,
                    "unit_of_measure": unit,
                    "is_active": active_val,
                }
            )

    created = 0
    updated = 0

    if items:
        conn = get_connection()
        cur = conn.cursor()
        try:
            for item in items:
                cur.execute(
                    "select product_id from products where product_code = %s",
                    (item["product_code"],),
                )
                row = cur.fetchone()
                if row:
                    cur.execute(
                        """
                        update products
                        set product_name = %s,
                            unit_of_measure = %s,
                            unit = %s,
                            is_active = %s,
                            active = %s,
                            updated_at = now()
                        where product_id = %s
                        """,
                        (
                            item["product_name"],
                            item["unit_of_measure"],
                            item["unit_of_measure"],
                            item["is_active"],
                            item["is_active"],
                            row[0],
                        ),
                    )
                    updated += 1
                else:
                    cur.execute(
                        """
                        insert into products
                            (product_code, product_name, unit_of_measure, is_active, unit, active)
                        values (%s, %s, %s, %s, %s, %s)
                        """,
                        (
                            item["product_code"],
                            item["product_name"],
                            item["unit_of_measure"],
                            item["is_active"],
                            item["unit_of_measure"],
                            item["is_active"],
                        ),
                    )
                    created += 1
            conn.commit()
        finally:
            cur.close()
            conn.close()

    return {
        "processed": len(items),
        "created_count": created,
        "updated_count": updated,
        "error_count": len(errors),
        "errors": errors,
    }

@app.get("/semi-finished")
def get_semi_finished():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        select
            semi_finished_id,
            semi_finished_code,
            semi_finished_name,
            unit,
            degas_days,
            active
        from semi_finished
        order by semi_finished_code
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    result = []
    for row in rows:
        result.append({
            "semi_finished_id": row[0],
            "semi_finished_code": row[1],
            "semi_finished_name": row[2],
            "unit": row[3],
            "unit_of_measure": row[3],
            "degas_days": row[4],
            "is_active": bool(row[5]),
        })

    return result


@app.get("/semi-finished/import-template")
def get_semi_finished_import_template():
    """
    Excel шаблон для импорта полуфабрикатов.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "semi_finished"
    headers = ["semi_finished_code", "semi_finished_name", "unit_of_measure", "degas_days", "is_active"]
    ws.append(headers)
    ws.append(["SF-EXAMPLE", "Пример полуфабриката", ALLOWED_UNITS[0], 0, "Да"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="semi_finished_import_template.xlsx"'},
    )


@app.post("/semi-finished/import")
def import_semi_finished(file: UploadFile = File(...)):
    """
    Импорт полуфабрикатов из Excel (.xlsx).
    """
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Требуется файл .xlsx")

    try:
        wb = load_workbook(file.file)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать Excel-файл")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Файл не содержит данных")

    header = [str(h).strip() if h else "" for h in rows[0]]
    expected_header = ["semi_finished_code", "semi_finished_name", "unit_of_measure", "degas_days", "is_active"]
    if [h.lower() for h in header] != expected_header:
        raise HTTPException(status_code=400, detail="Неверный заголовок файла")

    seen_codes = set()
    items = []
    errors = []

    def normalize_bool(val):
        if val is None:
            return None
        if isinstance(val, bool):
            return val
        s = str(val).strip().lower()
        if s in ["да", "yes", "true", "1"]:
            return True
        if s in ["нет", "no", "false", "0"]:
            return False
        return None

    for idx, row in enumerate(rows[1:], start=2):
        code_raw, name_raw, unit_raw, degas_raw, active_raw = row[:5]

        if all(
            (cell is None or (isinstance(cell, str) and cell.strip() == ""))
            for cell in (code_raw, name_raw, unit_raw, active_raw)
        ):
            continue

        row_errors = []
        code = (code_raw or "").strip()
        name = (name_raw or "").strip()

        try:
            unit = normalize_unit(unit_raw)
        except HTTPException as e:
            unit = None
            row_errors.append(e.detail)

        degas_days = 0
        if degas_raw not in (None, ""):
            try:
                degas_days = int(degas_raw)
                if degas_days < 0:
                    raise ValueError()
            except Exception:
                row_errors.append("degas_days должен быть неотрицательным целым")

        active = normalize_bool(active_raw)

        if not code:
            row_errors.append("не заполнен semi_finished_code")
        if code in seen_codes:
            row_errors.append("дублирующийся semi_finished_code в файле")
        if not name:
            row_errors.append("не заполнен semi_finished_name")
        if unit is None:
            row_errors.append("недопустимая единица измерения (разрешено: м² или м.п.)")
        if active is None:
            row_errors.append("недопустимое значение is_active (используйте Да/Нет)")

        if row_errors:
            errors.append({"row": idx, "errors": row_errors})
        else:
            seen_codes.add(code)
            items.append(
                {
                    "semi_finished_code": code,
                    "semi_finished_name": name,
                    "unit_of_measure": unit,
                    "degas_days": degas_days,
                    "is_active": active,
                }
            )

    created = 0
    updated = 0

    if items:
        conn = get_connection()
        cur = conn.cursor()
        try:
            for item in items:
                cur.execute(
                    "select semi_finished_id, degas_days from semi_finished where semi_finished_code = %s",
                    (item["semi_finished_code"],),
                )
                row = cur.fetchone()
                if row:
                    cur.execute(
                        """
                        update semi_finished
                        set semi_finished_name = %s,
                            unit = %s,
                            degas_days = %s,
                            active = %s,
                            updated_at = now()
                        where semi_finished_id = %s
                        """,
                        (
                            item["semi_finished_name"],
                            item["unit_of_measure"],
                            item["degas_days"],
                            item["is_active"],
                            row[0],
                        ),
                    )
                    updated += 1
                else:
                    # degas_days not provided, set 0 by default
                    cur.execute(
                        """
                        insert into semi_finished
                            (semi_finished_code, semi_finished_name, unit, degas_days, active)
                        values (%s, %s, %s, %s, %s)
                        """,
                        (
                            item["semi_finished_code"],
                            item["semi_finished_name"],
                            item["unit_of_measure"],
                            item["degas_days"],
                            item["is_active"],
                        ),
                    )
                    created += 1
            conn.commit()
        finally:
            cur.close()
            conn.close()

    return {
        "processed": len(items),
        "created_count": created,
        "updated_count": updated,
        "error_count": len(errors),
        "errors": errors,
    }

@app.put("/semi-finished/{semi_finished_code}")
def update_semi_finished(semi_finished_code: str, payload: SemiFinishedUpdate):
    update_fields = payload.dict(exclude_unset=True)
    if not update_fields:
        return {"status": "no_changes"}

    if "unit_of_measure" in update_fields:
        update_fields["unit_of_measure"] = normalize_unit(update_fields["unit_of_measure"])

    if "degas_days" in update_fields:
        try:
            if update_fields["degas_days"] is None:
                raise ValueError()
            update_fields["degas_days"] = int(update_fields["degas_days"])
            if update_fields["degas_days"] < 0:
                raise ValueError()
        except Exception:
            raise HTTPException(status_code=400, detail="Недопустимое значение degas_days (целое неотрицательное)")

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "select semi_finished_id from semi_finished where semi_finished_code = %s",
            (semi_finished_code,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Semi-finished not found")
        semi_finished_id = row[0]

        set_clauses = []
        params = []

        if "semi_finished_name" in update_fields:
            set_clauses.append("semi_finished_name = %s")
            params.append(update_fields["semi_finished_name"])

        if "unit_of_measure" in update_fields:
            set_clauses.append("unit = %s")
            params.append(update_fields["unit_of_measure"])

        if "is_active" in update_fields:
            set_clauses.append("active = %s")
            params.append(update_fields["is_active"])

        if "degas_days" in update_fields:
            set_clauses.append("degas_days = %s")
            params.append(update_fields["degas_days"])

        set_clauses.append("updated_at = now()")

        if len(params) == 0:
            return {"status": "no_changes"}

        sql = f"""
            update semi_finished
            set {', '.join(set_clauses)}
            where semi_finished_id = %s
            returning
                semi_finished_id,
                semi_finished_code,
                semi_finished_name,
                unit,
                degas_days,
                active
        """
        params.append(semi_finished_id)
        cur.execute(sql, tuple(params))
        updated = cur.fetchone()
        conn.commit()

        return {
            "semi_finished_id": updated[0],
            "semi_finished_code": updated[1],
            "semi_finished_name": updated[2],
            "unit": updated[3],
            "unit_of_measure": updated[3],
            "degas_days": updated[4],
            "is_active": bool(updated[5]),
        }
    finally:
        cur.close()
        conn.close()
@app.get("/routes")
def get_routes():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        select route_id, route_code, route_name
        from routes
        order by route_code
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    result = []
    for row in rows:
        result.append({
            "route_id": row[0],
            "route_code": row[1],
            "route_name": row[2],
        })

    return result

@app.get("/sales-plan")
def get_sales_plan():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        select
            sp.sales_plan_id,
            sp.period,
            p.product_code,
            p.product_name,
            sp.qty,
            sp.due_date
        from sales_plan sp
        join products p on p.product_id = sp.product_id
        order by sp.period, p.product_code
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    result = []
    for row in rows:
        result.append({
            "sales_plan_id": row[0],
            "period": row[1],
            "product_code": row[2],
            "product_name": row[3],
            "qty": float(row[4]),
            "due_date": str(row[5]) if row[5] else None,
        })

    return result


class SalesPlanItem(BaseModel):
    product_code: str
    period: str
    qty: float
    due_date: Optional[str] = None


@app.post("/sales-plan")
def save_sales_plan(items: List[SalesPlanItem]):
    """
    Bulk replace sales plan for provided periods.
    For each period present in payload, existing rows are deleted then reinserted.
    """
    if not items:
        return {"status": "empty", "inserted": 0}

    conn = get_connection()
    cur = conn.cursor()

    try:
        # delete existing rows for all periods present
        periods = {item.period for item in items}
        for period in periods:
            cur.execute("delete from sales_plan where period = %s", (period,))

        for item in items:
            cur.execute("select product_id from products where product_code = %s", (item.product_code,))
            product_row = cur.fetchone()
            if not product_row:
                raise HTTPException(status_code=400, detail=f"Unknown product_code: {item.product_code}")

            due_date_val = None
            if item.due_date:
                try:
                    due_date_val = datetime.date.fromisoformat(item.due_date)
                except ValueError:
                    raise HTTPException(status_code=400, detail=f"Invalid due_date: {item.due_date}")

            cur.execute(
                """
                insert into sales_plan (product_id, period, qty, due_date)
                values (%s, %s, %s, %s)
                """,
                (product_row[0], item.period, item.qty, due_date_val),
            )

        conn.commit()
        return {"status": "ok", "inserted": len(items), "periods": list(periods)}
    finally:
        cur.close()
        conn.close()
@app.get("/production-need")
def get_production_need():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        select
            pn.production_need_id,
            pn.period,
            p.product_code,
            p.product_name,
            pn.sales_plan_qty,
            pn.safety_stock_qty,
            pn.stock_qty,
            pn.required_qty
        from production_need pn
        join products p on p.product_id = pn.product_id
        order by pn.period, p.product_code
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    result = []
    for row in rows:
        result.append({
            "production_need_id": row[0],
            "period": row[1],
            "product_code": row[2],
            "product_name": row[3],
            "sales_plan_qty": float(row[4]),
            "safety_stock_qty": float(row[5]),
            "stock_qty": float(row[6]),
            "required_qty": float(row[7]),
        })

    return result

@app.get("/semi-finished-need")
def get_semi_finished_need():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        select
            sfn.semi_finished_need_id,
            sfn.period,
            p.product_code,
            p.product_name,
            sf.semi_finished_code,
            sf.semi_finished_name,
            sfn.required_product_qty,
            sfn.required_semi_finished_qty,
            sfn.source_component_id
        from semi_finished_need sfn
        join products p on p.product_id = sfn.source_product_id
        join semi_finished sf on sf.semi_finished_id = sfn.semi_finished_id
        order by sfn.period, p.product_code, sf.semi_finished_code
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    result = []
    for row in rows:
        result.append({
            "semi_finished_need_id": row[0],
            "period": row[1],
            "product_code": row[2],
            "product_name": row[3],
            "semi_finished_code": row[4],
            "semi_finished_name": row[5],
            "required_product_qty": float(row[6]),
            "required_semi_finished_qty": float(row[7]),
            "source_component_id": row[8],
        })

    return result

@app.get("/production-orders")
def get_production_orders():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        select
            po.order_id,
            r.route_code,
            po.step_no,
            sf_out.semi_finished_code,
            po.planned_qty,
            po.status,
            p.product_code
        from production_orders po
        join routes r on r.route_id = po.route_id
        join semi_finished sf_out on sf_out.semi_finished_id = po.output_semi_finished_id
        join semi_finished_need sfn on sfn.semi_finished_need_id = po.source_need_id
        join products p on p.product_id = sfn.source_product_id
        order by po.order_id
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    result = []
    for row in rows:
        result.append({
            "order_id": row[0],
            "route_code": row[1],
            "step_no": row[2],
            "output_semi_finished_code": row[3],
            "planned_qty": float(row[4]),
            "status": row[5],
            "product_code": row[6],
        })

    return result


@app.get("/products/{product_code}/structure")
def get_product_structure(product_code: str):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        select
            p.product_code,
            p.product_name,
            sf.semi_finished_code,
            sf.semi_finished_name,
            c.component_qty,
            c.product_qty,
            c.priority,
            r.route_code,
            r.route_name
        from products p
        join product_semi_finished_components c
            on c.product_id = p.product_id
        join semi_finished sf
            on sf.semi_finished_id = c.semi_finished_id
        join routes r
            on r.route_id = c.route_id
        where p.product_code = %s
        order by c.priority
    """, (product_code,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {"error": "Product not found"}

    result = {
        "product_code": rows[0][0],
        "product_name": rows[0][1],
        "components": []
    }

    for row in rows:
        result["components"].append({
            "semi_finished_code": row[2],
            "semi_finished_name": row[3],
            "component_qty": float(row[4]),
            "product_qty": float(row[5]),
            "priority": row[6],
            "route_code": row[7],
            "route_name": row[8],
        })

    return result
@app.get("/products/{product_code}/needs")
def get_product_needs(product_code: str):
    conn = get_connection()
    cur = conn.cursor()

    # 1. production_need
    cur.execute("""
        select
            pn.period,
            pn.required_qty
        from production_need pn
        join products p on p.product_id = pn.product_id
        where p.product_code = %s
    """, (product_code,))

    prod_need_row = cur.fetchone()

    # 2. semi_finished_need
    cur.execute("""
        select
            sf.semi_finished_code,
            sf.semi_finished_name,
            sfn.required_semi_finished_qty
        from semi_finished_need sfn
        join products p on p.product_id = sfn.source_product_id
        join semi_finished sf on sf.semi_finished_id = sfn.semi_finished_id
        where p.product_code = %s
        order by sf.semi_finished_code
    """, (product_code,))

    sf_rows = cur.fetchall()

    cur.close()
    conn.close()

    if not prod_need_row:
        return {"error": "Product not found or no needs"}

    result = {
        "product_code": product_code,
        "production_need": {
            "period": prod_need_row[0],
            "required_qty": float(prod_need_row[1])
        },
        "semi_finished_needs": []
    }

    for row in sf_rows:
        result["semi_finished_needs"].append({
            "semi_finished_code": row[0],
            "semi_finished_name": row[1],
            "required_qty": float(row[2])
        })

    return result
@app.get("/products/{product_code}/orders")
def get_product_orders(product_code: str):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        select
            po.order_id,
            r.route_code,
            r.route_name,
            po.step_no,
            sf.semi_finished_code,
            sf.semi_finished_name,
            po.planned_qty,
            po.status
        from production_orders po
        join routes r on r.route_id = po.route_id
        join semi_finished sf on sf.semi_finished_id = po.output_semi_finished_id
        join semi_finished_need sfn on sfn.semi_finished_need_id = po.source_need_id
        join products p on p.product_id = sfn.source_product_id
        where p.product_code = %s
        order by po.order_id
    """, (product_code,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {"error": "Product not found or no orders"}

    result = {
        "product_code": product_code,
        "orders": []
    }

    for row in rows:
        result["orders"].append({
            "order_id": row[0],
            "route_code": row[1],
            "route_name": row[2],
            "step_no": row[3],
            "output_semi_finished_code": row[4],
            "output_semi_finished_name": row[5],
            "planned_qty": float(row[6]),
            "status": row[7],
        })

    return result


@app.get("/products/{product_code}/actuals")
def get_product_actuals(product_code: str):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        select
            pa.actual_id,
            r.route_code,
            po.step_no,
            sf.semi_finished_code,
            sf.semi_finished_name,
            pa.fact_qty,
            pa.scrap_qty,
            pa.fact_start,
            pa.fact_finish
        from production_actuals pa
        join production_orders po on po.order_id = pa.order_id
        join routes r on r.route_id = po.route_id
        join semi_finished sf on sf.semi_finished_id = po.output_semi_finished_id
        join semi_finished_need sfn on sfn.semi_finished_need_id = po.source_need_id
        join products p on p.product_id = sfn.source_product_id
        where p.product_code = %s
        order by pa.actual_id
    """, (product_code,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    result = {
        "product_code": product_code,
        "actuals": []
    }

    for row in rows:
        result["actuals"].append({
            "actual_id": row[0],
            "route_code": row[1],
            "step_no": row[2],
            "output_semi_finished_code": row[3],
            "output_semi_finished_name": row[4],
            "fact_qty": float(row[5]),
            "scrap_qty": float(row[6]),
            "fact_start": str(row[7]) if row[7] else None,
            "fact_finish": str(row[8]) if row[8] else None,
        })

    return result
