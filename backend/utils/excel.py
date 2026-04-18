from typing import Iterable, Optional, Sequence

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook


_REQUIRE_XLSX_ERROR = "\u0422\u0440\u0435\u0431\u0443\u0435\u0442\u0441\u044f \u0444\u0430\u0439\u043b .xlsx"
_READ_XLSX_ERROR = "\u041d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u043f\u0440\u043e\u0447\u0438\u0442\u0430\u0442\u044c Excel-\u0444\u0430\u0439\u043b"
_EMPTY_XLSX_ERROR = "\u0424\u0430\u0439\u043b \u043d\u0435 \u0441\u043e\u0434\u0435\u0440\u0436\u0438\u0442 \u0434\u0430\u043d\u043d\u044b\u0445"
_INVALID_HEADER_ERROR = "\u041d\u0435\u0432\u0435\u0440\u043d\u044b\u0439 \u0437\u0430\u0433\u043e\u043b\u043e\u0432\u043e\u043a \u0444\u0430\u0439\u043b\u0430"

_BOOL_TRUE_VALUES = {"\u0434\u0430", "yes", "true", "1"}
_BOOL_FALSE_VALUES = {"\u043d\u0435\u0442", "no", "false", "0"}


def read_excel_rows(file: UploadFile, expected_header: Sequence[str]):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=_REQUIRE_XLSX_ERROR)

    try:
        workbook = load_workbook(file.file)
    except Exception:
        raise HTTPException(status_code=400, detail=_READ_XLSX_ERROR)

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail=_EMPTY_XLSX_ERROR)

    header = [str(value).strip() if value else "" for value in rows[0]]
    if [column.lower() for column in header] != list(expected_header):
        raise HTTPException(status_code=400, detail=_INVALID_HEADER_ERROR)

    return rows


def is_blank_excel_row(values: Iterable[object]) -> bool:
    return all(
        cell is None or (isinstance(cell, str) and cell.strip() == "")
        for cell in values
    )


def normalize_excel_bool(value) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value

    lowered = str(value).strip().lower()
    if lowered in _BOOL_TRUE_VALUES:
        return True
    if lowered in _BOOL_FALSE_VALUES:
        return False
    return None


def build_import_response(processed_count: int, created_count: int, updated_count: int, errors):
    return {
        "processed": processed_count,
        "created_count": created_count,
        "updated_count": updated_count,
        "error_count": len(errors),
        "errors": errors,
    }
